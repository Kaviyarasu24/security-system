import os
import io
from datetime import datetime

import pandas as pd


def format_entry_time(val):
	"""Format various entry_time values to DD:MM:YYYY  HH:MM:SS  AM/PM (12-hour)."""
	if val is None:
		return None
	if isinstance(val, datetime):
		return val.strftime("%d %B %Y  %I:%M:%S  %p")
	if isinstance(val, str):
		# try known input formats
		for fmt in ("%Y-%m-%d_%H-%M-%S", "%Y-%m-%d %H:%M:%S"):
			try:
				dt = datetime.strptime(val, fmt)
				return dt.strftime("%d %B %Y  %I:%M:%S  %p")
			except Exception:
				continue
		try:
			# try ISO parse
			dt = datetime.fromisoformat(val)
			return dt.strftime("%d %B %Y  %I:%M:%S  %p")
		except Exception:
			return val
	return val


def generate_reports(vehicle_records, csv_path="outputs/vehicle_report.csv", excel_path="outputs/vehicle_report.xlsx"):
	"""Generate CSV and Excel reports from vehicle_records.

	vehicle_records: dict keyed by id -> record dicts (as produced in main.py)
	Returns a dict with paths written.
	"""

	os.makedirs(os.path.dirname(csv_path), exist_ok=True)

	# Convert records to DataFrame
	records = list(vehicle_records.values())

	if not records:
		# create empty dataframe with expected columns
		df = pd.DataFrame(columns=[
			"id",
			"type",
			"plate",
			"entry_time",
			"exit_time",
			"dwell_time",
			"vehicle_image",
			"plate_image",
		])
	else:
		df = pd.DataFrame.from_records(records)

	# Ensure consistent column order
	cols = [
		"id",
		"type",
		"plate",
		"entry_time",
		"exit_time",
		"dwell_time",
		"vehicle_image",
		"plate_image",
	]

	for c in cols:
		if c not in df.columns:
			df[c] = None

	df = df[cols]

	# Normalize entry_time to requested format
	if "entry_time" in df.columns:
		df["entry_time"] = df["entry_time"].apply(format_entry_time)
	if "exit_time" in df.columns:
		df["exit_time"] = df["exit_time"].apply(format_entry_time)

	# Write CSV
	try:
		df.to_csv(csv_path, index=False)
	except Exception:
		# ignore CSV errors
		pass

	# Write Excel (simple sheet, without images)
	excel_written = None
	try:
		df.to_excel(excel_path, index=False)
		excel_written = excel_path
	except Exception:
		excel_written = None

	return {"csv": csv_path, "excel": excel_written}


def append_daily_excel(vehicle_records, date_str=None, excel_dir="reports"):
	"""Append vehicle_records to a daily Excel file named DD_MM_YYYY.xlsx, embedding images.

	Returns the path to the Excel file written, or None on failure.
	"""

	if date_str is None:
		date_str = datetime.now().strftime("%d_%m_%Y")

	os.makedirs(excel_dir, exist_ok=True)

	excel_path = os.path.join(excel_dir, f"{date_str}.xlsx")

	# Prepare dataframe from records
	records = list(vehicle_records.values())

	if not records:
		df_new = pd.DataFrame(columns=[
			"id",
			"type",
			"plate",
			"entry_time",
			"exit_time",
			"dwell_time",
			"vehicle_image",
			"plate_image",
		])
	else:
		df_new = pd.DataFrame.from_records(records)

	# Ensure consistent columns
	cols = [
		"id",
		"type",
		"plate",
		"entry_time",
		"exit_time",
		"dwell_time",
		"vehicle_image",
		"plate_image",
	]

	for c in cols:
		if c not in df_new.columns:
			df_new[c] = None

	df_new = df_new[cols]

	# Normalize entry_time for new records as well
	if "entry_time" in df_new.columns:
		df_new["entry_time"] = df_new["entry_time"].apply(format_entry_time)
	if "exit_time" in df_new.columns:
		df_new["exit_time"] = df_new["exit_time"].apply(format_entry_time)

	try:
		# Lazy imports for optional dependencies
		from openpyxl import load_workbook, Workbook
		from openpyxl.drawing.image import Image as XLImage
		from openpyxl.utils import get_column_letter
		from PIL import Image as PILImage

		if os.path.exists(excel_path):
			try:
				wb = load_workbook(excel_path)
				ws = wb.active
				start_row = ws.max_row + 1
			except Exception:
				wb = Workbook()
				ws = wb.active
				# write header
				for idx, h in enumerate(cols, start=1):
					ws.cell(row=1, column=idx, value=h)
				start_row = 2
		else:
			wb = Workbook()
			ws = wb.active
			# write header
			for idx, h in enumerate(cols, start=1):
				ws.cell(row=1, column=idx, value=h)

			start_row = 2

		def insert_image(path, row, col_index):
			if not path:
				return
			if not os.path.exists(path):
				return
			try:
				pil_img = PILImage.open(path)
				# Resize to reasonable thumbnail
				max_w, max_h = 320, 180
				w, h = pil_img.size
				scale = min(max_w / w, max_h / h, 1.0)
				new_w = int(w * scale)
				new_h = int(h * scale)
				if scale < 1.0:
					pil_img = pil_img.resize((new_w, new_h), PILImage.LANCZOS)

				bio = io.BytesIO()
				pil_img.save(bio, format="PNG")
				bio.seek(0)

				img = XLImage(bio)
				anchor = ws.cell(row=row, column=col_index).coordinate
				ws.add_image(img, anchor)

				# Try to fit the image inside the cell by adjusting column width and row height
				try:
					col_letter = get_column_letter(col_index)
					# Approximate conversion from pixels to Excel column width
					# This is heuristic: width units ~= pixels * 0.14
					desired_width = new_w * 0.14 + 2
					current_width = ws.column_dimensions[col_letter].width or 0
					if desired_width > current_width:
						ws.column_dimensions[col_letter].width = desired_width

					# Row height in points (approx): set proportional to image height
					desired_height = max(40, int(new_h * 0.75))
					current_height = ws.row_dimensions[row].height or 0
					if desired_height > current_height:
						ws.row_dimensions[row].height = desired_height
				except Exception:
					# ignore column/row sizing failures
					pass
			except Exception:
				return

		# Append rows with embedded images
		row_idx = start_row
		for _, rec in df_new.iterrows():
			ws.cell(row=row_idx, column=1, value=rec.get("id"))
			ws.cell(row=row_idx, column=2, value=rec.get("type"))
			ws.cell(row=row_idx, column=3, value=rec.get("plate"))
			ws.cell(row=row_idx, column=4, value=rec.get("entry_time"))
			ws.cell(row=row_idx, column=5, value=rec.get("exit_time"))
			ws.cell(row=row_idx, column=6, value=rec.get("dwell_time"))

			# embed images in columns 7 and 8
			insert_image(rec.get("vehicle_image"), row_idx, 7)
			insert_image(rec.get("plate_image"),   row_idx, 8)

			row_idx += 1

		wb.save(excel_path)
		return excel_path

	except PermissionError:
		# File may be open in Excel
		return None
	except Exception:
		return None
